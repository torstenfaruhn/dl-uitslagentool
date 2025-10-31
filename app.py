from flask import Flask, render_template, request, Response, abort, send_file
from converter_regiosport import excel_to_txt_regiosport
from converter_amateur import excel_to_txt_amateur

# Voor sjablonen
from openpyxl import Workbook
import io

app = Flask(__name__)


# -----------------------------
# UI
# -----------------------------
@app.get("/")
def index():
    return render_template("index.html")


# -----------------------------
# Convert endpoints
# -----------------------------
@app.post("/convert/regiosport")
def convert_regiosport():
    file = request.files.get("file_regio")
    if not file or file.filename == "":
        return abort(400, "Geen bestand geüpload (Regiosport).")
    try:
        txt = excel_to_txt_regiosport(file.read())
    except Exception as e:
        return abort(400, f"Kon Regiosport-bestand niet verwerken: {e}")
    return Response(
        txt,
        mimetype="text/plain; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=cue_export_regiosport.txt"},
    )


@app.post("/convert/amateur")
def convert_amateur():
    file = request.files.get("file_amateur")
    if not file or file.filename == "":
        return abort(400, "Geen bestand geüpload (Amateurvoetbal).")
    try:
        txt = excel_to_txt_amateur(file.read())
    except Exception as e:
        return abort(400, f"Kon Amateur-bestand niet verwerken: {e}")
    return Response(
        txt,
        mimetype="text/plain; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=cue_export_amateur.txt"},
    )


# -----------------------------
# Template (leeg invoerdocument) endpoints
# -----------------------------
def _xls_bytes_from_workbook(wb: Workbook) -> bytes:
    """Helper: zet Workbook om naar bytes voor download."""
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


@app.get("/template/amateur")
def template_amateur():
    """
    Leeg sjabloon voor DL amateurvoetbal tool.
    Kolommen in één sheet:
      Thuisclub | Uitclub | ThuisDoelpunten | UitDoelpunten | RustThuis | RustUit | Doelpuntenmakers
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Amateurvoetbal"
    ws.append(
        [
            "Thuisclub",
            "Uitclub",
            "ThuisDoelpunten",
            "UitDoelpunten",
            "RustThuis",
            "RustUit",
            "Doelpuntenmakers",
        ]
    )
    data = _xls_bytes_from_workbook(wb)
    return send_file(
        io.BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="DL_amateurvoetbal_invoerdocument.xlsx",
    )


@app.get("/template/regiosport")
def template_regiosport():
    """
    Leeg sjabloon voor DL regiosport tool.

    Sheet 1: 'Sporten met uitslagregel'
      Kolommen: Label | Waarde
      Gebruik o.a.: SPORT, EVENEMENT, UITSLAGREGEL 1..N. Lege rij = nieuw blok.

    Sheet 2: 'Sporten met stand'
      Kolommen: Label | Thuis | HS | Uit | AS
      Optionele kopregel (Label leeg) kan gebruikt worden voor kolomtitels.
      'STAND' in Label sluit een blok af en plaatst de standtekst.
    """
    wb = Workbook()

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Sporten met uitslagregel"
    ws1.append(["Label", "Waarde"])
    # (Voorbeeldregels eventueel later toevoegen)

    # Sheet 2
    ws2 = wb.create_sheet("Sporten met stand")
    ws2.append(["Label", "Thuis", "HS", "Uit", "AS"])

    data = _xls_bytes_from_workbook(wb)
    return send_file(
        io.BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="DL_regiosport_invoerdocument.xlsx",
    )


# -----------------------------
# Main
# -----------------------------
if __name__ == "__main__":
    # Voor lokaal testen:
    # python app.py -> http://localhost:8000
    app.run(host="0.0.0.0", port=8000, debug=False)
